#!/usr/bin/env python3
"""Refresh data/team_draft_2026.csv from sv-org-review's master xlsx.

Reads the two tabs we care about:
  - "Pool Amounts (26)": one row per team in the form "<Team Name>: $X,XXX,XXX".
  - "Slot Values (26)":  numbered pick rows in the form "<n>. <Team Name>: $X,XXX,XXX",
    organized into round sections (First round / CBA / Second round / CBB / Third round
    / Fourth round / ...). We keep picks from rounds 1-3 (incl. Comp Balance A & B and
    any Prospect Promotion / Free-agent Compensation picks that slot into that range).

Run after the sv-org-review xlsx is updated:
    python3 scripts/refresh_team_draft.py
    # or point at a different file:
    python3 scripts/refresh_team_draft.py --xlsx /path/to/Org.Review.2026.update_X.xlsx

Output: data/team_draft_2026.csv with columns abbrev, pool_amount, picks.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys

DEFAULT_XLSX = os.path.expanduser(
    '~/Desktop/claude/sv-org-review/Org.Review.2026.update_5-13-26.xlsx'
)

# Short team name (as it appears in the two source tabs) -> 3-letter abbrev used
# everywhere else in sv-teamintel. "Athletics" → "ATH" (CLAUDE.md note: this repo
# uses ATH, not OAK, since the city was dropped in 2025).
NAME_TO_ABBREV = {
    'Diamondbacks': 'ARI', 'Braves': 'ATL', 'Orioles': 'BAL', 'Red Sox': 'BOS',
    'Cubs': 'CHC', 'White Sox': 'CHW', 'Reds': 'CIN', 'Guardians': 'CLE',
    'Rockies': 'COL', 'Tigers': 'DET', 'Astros': 'HOU', 'Royals': 'KC',
    'Angels': 'LAA', 'Dodgers': 'LAD', 'Marlins': 'MIA', 'Brewers': 'MIL',
    'Twins': 'MIN', 'Mets': 'NYM', 'Yankees': 'NYY', 'Athletics': 'ATH',
    'Phillies': 'PHI', 'Pirates': 'PIT', 'Padres': 'SD', 'Giants': 'SF',
    'Mariners': 'SEA', 'Cardinals': 'STL', 'Rays': 'TB', 'Rangers': 'TEX',
    'Blue Jays': 'TOR', 'Nationals': 'WSH',
}

# Round headers in Slot Values (26) that mark the start of a numbered round.
# We track the active round explicitly; sub-section headers like "Competitive
# Balance Round A/B", "Prospect Promotion Incentive Pick", and "Free-agent
# Compensation Pick" don't change the round — they slot inside whatever round
# we just entered. Tracking round number directly avoids a bug where a
# Free-agent Compensation Pick header between Round 4 and Round 5 would
# incorrectly re-enable round-1-3 scope.
ROUND_HEADERS = {
    'First round': 1, 'Second round': 2, 'Third round': 3, 'Fourth round': 4,
    'Fifth round': 5, 'Sixth round': 6, 'Seventh round': 7, 'Eighth round': 8,
    'Ninth Round': 9, '10th round': 10,
}
MAX_ROUND = 3

PICK_RE = re.compile(r'^\s*(\d+)\.\s+(.+?):\s*\$')
POOL_RE = re.compile(r'^(.+?):\s*\$([\d,]+)\s*$')


def fmt_pool_millions(dollar_str: str) -> str:
    """'19,130,700' -> '$19.13m' (matches the existing CSV format)."""
    n = int(dollar_str.replace(',', ''))
    return f'${n / 1_000_000:.2f}m'


def extract(xlsx_path: str):
    import openpyxl  # local import so the script only needs openpyxl when invoked
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    pool_ws = wb['Pool Amounts (26)']
    slot_ws = wb['Slot Values (26)']

    pools: dict[str, str] = {}
    for row in pool_ws.iter_rows(min_row=1, values_only=True):
        cell = row[0]
        if not cell:
            continue
        m = POOL_RE.match(str(cell).strip())
        if not m:
            continue
        name, dollars = m.group(1).strip(), m.group(2)
        abbrev = NAME_TO_ABBREV.get(name)
        if not abbrev:
            print(f'WARN: unknown team in Pool Amounts: {name!r}', file=sys.stderr)
            continue
        pools[abbrev] = fmt_pool_millions(dollars)

    picks: dict[str, list[int]] = {}
    round_num = 0  # 0 = haven't entered Round 1 yet
    for row in slot_ws.iter_rows(min_row=1, values_only=True):
        cell = row[0]
        if cell is None:
            continue
        text = str(cell).strip()
        new_round = ROUND_HEADERS.get(text)
        if new_round is not None:
            round_num = new_round
            continue
        if round_num == 0 or round_num > MAX_ROUND:
            continue
        m = PICK_RE.match(text)
        if not m:
            continue
        pick_num, name = int(m.group(1)), m.group(2).strip()
        abbrev = NAME_TO_ABBREV.get(name)
        if not abbrev:
            print(f'WARN: unknown team in Slot Values: {name!r} (pick {pick_num})', file=sys.stderr)
            continue
        picks.setdefault(abbrev, []).append(pick_num)

    return pools, picks


def write_csv(out_path: str, pools: dict[str, str], picks: dict[str, list[int]]):
    abbrevs = sorted(set(pools) | set(picks))
    with open(out_path, 'w', newline='') as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(['abbrev', 'pool_amount', 'picks'])
        for ab in abbrevs:
            picks_str = ', '.join(str(n) for n in picks.get(ab, []))
            w.writerow([ab, pools.get(ab, ''), picks_str])


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--xlsx', default=DEFAULT_XLSX,
                   help=f'Path to Org Review xlsx (default: {DEFAULT_XLSX})')
    p.add_argument('--out', default=os.path.join(os.path.dirname(__file__), '..', 'data', 'team_draft_2026.csv'))
    args = p.parse_args()
    if not os.path.exists(args.xlsx):
        sys.exit(f'ERROR: xlsx not found at {args.xlsx}')
    pools, picks = extract(args.xlsx)
    out = os.path.abspath(args.out)
    write_csv(out, pools, picks)
    n = max(len(pools), len(picks))
    print(f'Wrote {n} teams to {out}')
    print(f'  pools: {len(pools)}   picks: {len(picks)}')
    missing_pool  = [a for a in NAME_TO_ABBREV.values() if a not in pools]
    missing_picks = [a for a in NAME_TO_ABBREV.values() if a not in picks]
    if missing_pool:
        print(f'  WARN: teams without pool data: {missing_pool}')
    if missing_picks:
        print(f'  WARN: teams without picks data: {missing_picks}')


if __name__ == '__main__':
    main()
